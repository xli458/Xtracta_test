# -*- coding: utf-8 -*-
"""
Created on Wed Nov 16 09:37:22 2016

This program is to find the supplier name of that invoice by matching the given list of supplier names to the invoice.

@author: xli458
"""

import math
import time
import openpyxl

#import string
#punctuations = list(string.punctuation) # get all punctuations

# Read in the files
with open('invoice.txt') as f:
    invoice_r = f.readlines()

with open('suppliernames_mod.txt') as f:
    Suppliers_R = f.readlines()
   
start_time = time.clock()  # start the time recording

# First store all suppliers into a list
Supplier_list = []
for Sup_tmp in Suppliers_R:
    rw_tmp = Sup_tmp.split(',')
    if rw_tmp[1][-1:] == '\n':
        rw_tmp[1] = rw_tmp[1][:-1].lower()
    Supplier_list.append(rw_tmp[1])   # only insert the name of the suppliers
# sort the names alphabetically
Supplier_list= sorted(Supplier_list)

Sup_split_name = []
for Sup_tmp in Supplier_list:    # split the  names
    rw_tmp = Sup_tmp.split(' ')    # split through ' '
    Sup_split_name.append(rw_tmp)
Supply_num = len(Supplier_list)   # number of suppliers

# take out the key information from the invoice obtained from OCR
word_list = []
pos_max = -1    # maximum position
line_max = -1   # max line number
page_max = -1   # max page number
for ino_tmp in invoice_r:
    rw_tmp = ino_tmp.split('\'')   # note that comma could be contained in the "word"
    # Take out the word 
    word_tmp = rw_tmp[11].lower()  # change to lower case
    # This is the position of word at the line  
    pos_id = int(rw_tmp[2][2:-2]) 
    if pos_id > pos_max:
        pos_max = pos_id
    # Take out the line id
    line_id = int(rw_tmp[14][2:-2])
    if line_id > line_max:
        line_max = line_id
    # Take page id 
    page_id = int(rw_tmp[24][2:-2])
    if page_id > page_max:
        page_max = page_id
    word_list.append([page_id, line_id, pos_id, word_tmp])
    
# now reconstruct the invoice according to page, line, position information
# initialize high dimensional array  (reconstruct the invoice)
invoice_rec = [[['' for k in range(pos_max+1)] for j in range(line_max+1)] for i in range(page_max)]  
for rw_tmp in word_list:
    invoice_rec[rw_tmp[0]-1][rw_tmp[1]][rw_tmp[2]] = rw_tmp[3] 

Sup_name_found = []
# now search through the reconstructed invoice 
for page in invoice_rec:
    for line in page:
        line_pt = -1  # position of the word in line
        for word_tmp in line:
            line_pt += 1   # index of the word_tmp
            
            indx = -1
            start_pt = 0
            end_pt = len(Sup_split_name)
            
            while(end_pt > start_pt+1):
                mid = math.ceil((end_pt + start_pt) / 2) 
                    
                if Sup_split_name[mid][0] == word_tmp:
                    indx = mid
                    name_len = len(Sup_split_name[indx])  # length of the company name
                    if line_pt + name_len < line_max:
                        for i in range(0, name_len): 
                            if (line[line_pt + i] != Sup_split_name[indx][i]):
                                indx = -1  # not found
                                break
                        
                        if indx < 0:
                            if (line[line_pt + i] < Sup_split_name[mid][i]):
                                end_pt = mid
                            if (line[line_pt + i] > Sup_split_name[mid][i]):
                                start_pt = mid
                        if indx >= 0:
                            break
                        
                elif Sup_split_name[mid][0] > word_tmp:
                    end_pt = mid
                elif Sup_split_name[mid][0] < word_tmp: 
                    start_pt = mid
 
            if end_pt == start_pt+1 and indx <0:
                if Sup_split_name[end_pt][0] == word_tmp:
                    indx = end_pt
                    name_len = len(Sup_split_name[indx])  # length of the company name
                    if line_pt + name_len < line_max:
                        for i in range(0, name_len): 
                            if (line[line_pt + i] != Sup_split_name[indx][i]):
                                indx = -1  # not found
                                break

                if Sup_split_name[start_pt][0] == word_tmp:    
                    indx = start_pt                    
                    name_len = len(Sup_split_name[indx])  # length of the company name
                    if line_pt + name_len < line_max:
                        for i in range(0, name_len): 
                            if (line[line_pt + i] != Sup_split_name[indx][i]):
                                print(line[line_pt + i])
                                print(Sup_split_name[indx][i])
                                indx = -1  # not found
                                break
                    
            if indx >= 0:
                Sup_name_found.append(Supplier_list[indx])

print("--- %s seconds ---" % (time.clock() - start_time))   

# file writing
wb = openpyxl.Workbook()
# wb.get_sheet_names()
sheet_dic = wb.get_sheet_by_name('Sheet')
for i in range(len(word_list)):
    for j in range(len(word_list[i])):
        sheet_dic.cell(row = i, column = j).value = word_list[i][j] 
wb.save('word_list.xlsx')

wb = openpyxl.Workbook()
# wb.get_sheet_names()
sheet_dic = wb.get_sheet_by_name('Sheet')
for i in range(len(invoice_rec[0])):
    for j in range(len(invoice_rec[0][i])):
        sheet_dic.cell(row = i, column = j).value = invoice_rec[0][i][j] 
wb.save('invoice_rec.xlsx')

wb = openpyxl.Workbook()
# wb.get_sheet_names()
sheet_dic = wb.get_sheet_by_name('Sheet')
for i in range(len(Sup_name_found)):
        sheet_dic.cell(row = i, column = 0).value = Sup_name_found[0]
wb.save('Sup_name_found.xlsx')